from openpyxl import load_workbook

from app.db.database import SessionLocal
from app.db.models import PatientORM, UKSVisitORM

EXCEL_FILE = r"C:\Users\indah\OneDrive\Dokumen\New project\laporan_kunjungan.xlsx"

db = SessionLocal()

wb = load_workbook(EXCEL_FILE)
ws = wb.active

patients = db.query(PatientORM).all()

success = 0
skipped = []

for row in ws.iter_rows(min_row=2, values_only=True):

    visit_date = row[1]
    excel_name = str(row[2]).strip()

    complaint = str(row[5] or "").strip()
    diagnosis = str(row[6] or "").strip()
    examination = str(row[7] or "").strip()
    treatment = str(row[8] or "").strip()

    patient = None

    # exact match dulu
    for p in patients:
        if p.name.lower() == excel_name.lower():
            patient = p
            break

    # partial match
    if not patient:
        for p in patients:
            if excel_name.lower() in p.name.lower():
                patient = p
                break

    if not patient:
        skipped.append(excel_name)
        continue

    visit = UKSVisitORM(
        patient_id=patient.id,
        visit_date=visit_date,
        complaint=complaint,
        diagnosis=diagnosis,
        examination=examination,
        treatment=treatment,
        notes="Imported from Excel"
    )

    db.add(visit)
    success += 1

db.commit()

print(f"\nSUCCESS IMPORT: {success}")

print("\nSKIPPED:")
for name in sorted(set(skipped)):
    print("-", name)
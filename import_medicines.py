from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from app.db.database import SessionLocal
from app.db.models import MedicineInventoryORM


def _norm(value: object) -> str:
    text = str(value or "").strip().lower()
    for ch in [" ", "_", "-", "/", "\\", ".", ",", ":", ";", "(", ")"]:
        text = text.replace(ch, "")
    return text


def _pick(headers: dict[str, int], aliases: list[str]) -> int | None:
    for alias in aliases:
        if alias in headers:
            return headers[alias]
    return None


def _find_header_row(rows: list[tuple]) -> tuple[int, dict[str, int]] | tuple[None, dict[str, int]]:
    name_aliases = {"namaobat", "obat", "nama", "namabarang", "barang", "medicine", "medicinename"}
    stock_aliases = {"stok", "stock", "qty", "jumlah", "saldo", "sisa", "stokakhir", "saldoakhir"}

    max_scan = min(len(rows), 25)
    for row_idx in range(max_scan):
        header_map: dict[str, int] = {}
        for idx, cell in enumerate(rows[row_idx]):
            key = _norm(cell)
            if key:
                header_map[key] = idx

        has_name = any(alias in header_map for alias in name_aliases)
        has_stock = any(alias in header_map for alias in stock_aliases)
        if has_name and has_stock:
            return row_idx, header_map

    return None, {}


def run_import(file_path: Path, sheet_name: str | None = None) -> None:
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("Sheet kosong.")
        return

    header_idx, header_map = _find_header_row(rows)
    if header_idx is None:
        print("Kolom wajib tidak ditemukan. Header terdeteksi tidak cocok.")
        print("Tip: pakai kolom seperti 'Nama Obat' dan 'Stok' atau kirim nama sheet yang benar.")
        print("Daftar sheet:", ", ".join(wb.sheetnames))
        return

    name_col = _pick(header_map, ["namaobat", "obat", "nama", "namabarang", "barang", "medicine", "medicinename"])
    stock_col = _pick(header_map, ["stok", "stock", "qty", "jumlah", "saldo", "sisa", "stokakhir", "saldoakhir"])
    unit_col = _pick(header_map, ["satuan", "unit"])
    min_col = _pick(header_map, ["stokminimum", "minimum", "minstock", "minimumstock"])

    if name_col is None or stock_col is None:
        print("Kolom wajib tidak ditemukan. Minimal harus ada 'Nama Obat' dan 'Stok'.")
        return

    db = SessionLocal()
    created = 0
    updated = 0
    skipped = 0

    try:
        for row in rows[header_idx + 1:]:
            name_raw = row[name_col] if name_col < len(row) else None
            stock_raw = row[stock_col] if stock_col < len(row) else None
            if name_raw is None or stock_raw is None:
                skipped += 1
                continue

            name = str(name_raw).strip()
            if not name:
                skipped += 1
                continue

            try:
                stock = int(float(str(stock_raw).replace(",", ".").strip()))
            except Exception:
                skipped += 1
                continue

            unit = "tablet"
            if unit_col is not None and unit_col < len(row) and row[unit_col] is not None:
                unit = str(row[unit_col]).strip() or "tablet"

            minimum_stock = 10
            if min_col is not None and min_col < len(row) and row[min_col] is not None:
                try:
                    minimum_stock = int(float(str(row[min_col]).replace(",", ".").strip()))
                except Exception:
                    minimum_stock = 10

            existing = (
                db.query(MedicineInventoryORM)
                .filter(MedicineInventoryORM.name.ilike(name))
                .first()
            )
            if existing:
                existing.stock = stock
                existing.unit = unit
                existing.minimum_stock = minimum_stock
                db.add(existing)
                updated += 1
            else:
                db.add(
                    MedicineInventoryORM(
                        name=name,
                        unit=unit,
                        stock=max(stock, 0),
                        minimum_stock=max(minimum_stock, 0),
                    )
                )
                created += 1

        db.commit()
    finally:
        db.close()

    print(
        f"Import selesai. Ditambah: {created}, diupdate: {updated}, dilewati: {skipped}."
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Import stok obat UKS dari file Excel.")
    parser.add_argument("file", help="Path file .xlsx")
    parser.add_argument("--sheet", default=None, help="Nama sheet (opsional)")
    args = parser.parse_args()

    file_path = Path(args.file)
    if not file_path.exists():
        raise FileNotFoundError(f"File tidak ditemukan: {file_path}")

    run_import(file_path=file_path, sheet_name=args.sheet)


if __name__ == "__main__":
    main()

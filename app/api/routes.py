from datetime import datetime, timedelta
from html import escape
import base64
import json
import os
from pathlib import Path
import shutil
import requests
from dotenv import load_dotenv
try:
    from openai import OpenAI
except ImportError:  # pragma: no cover
    OpenAI = None
load_dotenv()
from io import BytesIO
from app.api.recommendations import (
    letterhead_flowable,
    signature_image_flowable,
    qr_code_flowable,
)

from fastapi import APIRouter, Depends, HTTPException, Query, Request, Response, status
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.auth.dependencies import get_current_user, require_roles
from app.auth.security import (
    create_access_token,
    get_access_token_expire_seconds,
    hash_password,
    verify_password,
)
from app.core.expert_system import NursingExpertSystem
from app.db.dependencies import get_db
from app.db.models import (
    AssessmentORM,
    AuditLogORM,
    CKGStudentORM,
    MedicineInventoryORM,
    PatientORM,
    RecommendationORM,
    RecommendationLetterORM,
    UKSMedicationORM,
    UKSVisitORM,
    UserORM,
    MedicineTransactionORM,
)
from app.models.schemas import (
    AICareSuggestionRequest,
    AICareSuggestionResponse,
    AssessmentResponse,
    AssessmentSummary,
    ChangePasswordRequest,
    ComplaintStat,
    LoginRequest,
    MedicineInventoryCreate,
    AuditLogListResponse,
    AuditLogResponse,
    MedicineInventoryResponse,
    MedicineStockAdjustment,
    MedicineInventoryUpdate,
    NursingAssessment,
    Patient,
    PatientCreate,
    PatientAssessmentsResponse,
    PatientSummary,
    PasswordResetRequest,
    TokenResponse,
    UKSDailyReportResponse,
    UKSMedicationCreate,
    UKSMedicationResponse,
    UKSMonthlyReportResponse,
    UKSReferralUpdate,
    UKSVisitCreate,
    UKSVisitResponse,
    UserCreate,
    UserProfileUpdate,
    UserUpdate,
    UserResponse,
)

router = APIRouter(prefix="/api", tags=["EMR Keperawatan"])
expert_system = NursingExpertSystem()
client = (
    OpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=os.getenv("OPENROUTER_API_KEY"),
        timeout=float(os.getenv("OPENROUTER_TIMEOUT_SECONDS", "8")),
        max_retries=0,
    )
    if OpenAI is not None
    else None
)
OPENROUTER_MODEL = os.getenv("OPENROUTER_MODEL", "openai/gpt-oss-120b:free")

ROLE_ADMIN = "admin"
ROLE_PERAWAT = "perawat"


def build_local_care_suggestion(complaint: str, examination: str) -> str:
    text = f"{complaint or ''} {examination or ''}".lower()
    findings: list[tuple[str, str, str, str]] = []

    if any(word in text for word in ("gigi", "gusi", "karies", "sariawan", "mulut")):
        findings.append(
            (
                "Nyeri akut terkait gangguan pada area gigi dan mulut.",
                "Kaji lokasi dan skala nyeri, periksa pembengkakan/kemerahan, anjurkan kebersihan mulut, dan berikan analgesik sesuai protokol UKS.",
                "Pantau respons nyeri 15-30 menit, catat obat yang diberikan, serta edukasi siswa untuk menghindari makanan terlalu keras/manis sementara.",
                "Rujuk ke fasilitas kesehatan bila nyeri menetap, ada bengkak, demam, perdarahan, atau siswa sulit makan.",
            )
        )
    if any(word in text for word in ("batuk", "sesak", "napas", "pilek", "flu", "tenggorokan")):
        findings.append(
            (
                "Bersihan jalan napas tidak efektif terkait iritasi saluran napas.",
                "Pantau frekuensi napas dan suhu, anjurkan minum air hangat, posisikan nyaman, dan ajarkan teknik batuk efektif.",
                "Observasi bunyi napas, kemampuan mengeluarkan dahak, dan respons setelah istirahat; gunakan masker bila batuk aktif.",
                "Hubungi wali asuh atau rujuk bila sesak, demam tinggi, napas cepat, atau keluhan tidak membaik.",
            )
        )
    if any(word in text for word in ("pusing", "sakit kepala", "lemas", "mual", "nyeri ulu hati", "perut")):
        findings.append(
            (
                "Gangguan kenyamanan akut terkait keluhan pusing, lemas, atau nyeri abdomen.",
                "Istirahatkan siswa di ruang UKS, pantau tanda vital, kaji pola makan/minum terakhir, dan berikan cairan oral bila tidak mual berat.",
                "Evaluasi skala keluhan setelah 15-30 menit, batasi aktivitas fisik, dan dokumentasikan faktor pemicu yang ditemukan.",
                "Rujuk atau hubungi wali asuh bila keluhan memberat, muntah berulang, nyeri perut hebat, atau pusing disertai tanda bahaya.",
            )
        )
    if any(word in text for word in ("luka", "jatuh", "memar", "terkilir", "benturan", "berdarah")):
        findings.append(
            (
                "Risiko infeksi atau nyeri akut terkait cedera jaringan.",
                "Bersihkan luka sesuai prosedur, tekan perdarahan ringan, kompres area memar, dan imobilisasi sementara bila dicurigai terkilir.",
                "Catat lokasi luka, ukuran, nyeri, dan kemampuan gerak; pantau tanda infeksi atau pembengkakan bertambah.",
                "Rujuk bila luka dalam, perdarahan sulit berhenti, deformitas, nyeri berat, atau keterbatasan gerak signifikan.",
            )
        )

    if not findings:
        findings.append(
            (
                "Gangguan kenyamanan akut terkait keluhan fisik siswa.",
                "Observasi keadaan umum dan tanda vital, anjurkan istirahat, berikan cairan oral sesuai kondisi, dan lakukan edukasi singkat sesuai keluhan.",
                "Pantau respons siswa 15-30 menit, dokumentasikan perubahan kondisi, serta pastikan siswa tidak kembali beraktivitas berat terlalu cepat.",
                "Evaluasi ulang sesuai kondisi; hubungi wali asuh bila keluhan berulang, menetap, atau muncul tanda bahaya.",
            )
        )

    diagnoses = " ".join(item[0] for item in findings[:2])
    interventions = " ".join(item[1] for item in findings[:2])
    notes = " ".join(f"Implementasi: {item[2]} Tindak lanjut: {item[3]}" for item in findings[:2])
    return (
        "Diagnosa Keperawatan:\n"
        f"{diagnoses}\n\n"
        "Tindakan Keperawatan:\n"
        f"{interventions}\n\n"
        "Catatan:\n"
        f"{notes}"
    )


def get_openrouter_models() -> list[str]:
    raw_models = os.getenv("OPENROUTER_MODELS") or OPENROUTER_MODEL
    models = [model.strip() for model in raw_models.split(",") if model.strip()]
    if OPENROUTER_MODEL and OPENROUTER_MODEL not in models:
        models.insert(0, OPENROUTER_MODEL)
    return models or ["openai/gpt-oss-120b:free"]


def _user_response(user: UserORM) -> UserResponse:
    return UserResponse(
        id=user.id,
        username=user.username,
        full_name=user.full_name,
        role=user.role,
        is_active=user.is_active,
        nip=getattr(user, "nip", None),
        jabatan=getattr(user, "jabatan", None),
        signature_image=getattr(user, "signature_image", None),
        created_at=getattr(user, "created_at", None),
        updated_at=getattr(user, "updated_at", None),
    )


def write_audit_log(
    db: Session,
    user: UserORM | None,
    action: str,
    entity_type: str,
    entity_id: str | int | None = None,
    details: str | None = None,
) -> None:
    db.add(
        AuditLogORM(
            user_id=user.id if user else None,
            username=user.username if user else None,
            action=action,
            entity_type=entity_type,
            entity_id=str(entity_id) if entity_id is not None else None,
            details=details,
        )
    )


def _active_admin_count(db: Session) -> int:
    return (
        db.query(UserORM)
        .filter(UserORM.role == ROLE_ADMIN, UserORM.is_active.is_(True))
        .count()
    )


def _build_top_complaints(visits: list[UKSVisitORM], limit: int = 5) -> list[ComplaintStat]:
    counts: dict[str, int] = {}
    for visit in visits:
        key = visit.complaint.strip()
        counts[key] = counts.get(key, 0) + 1
    sorted_items = sorted(counts.items(), key=lambda item: (-item[1], item[0]))
    return [ComplaintStat(complaint=complaint, total=total) for complaint, total in sorted_items[:limit]]


def _validate_date_yyyy_mm_dd(value: str) -> str:
    try:
        datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Date format must be YYYY-MM-DD")
    return value


def _validate_month_yyyy_mm(value: str) -> str:
    try:
        datetime.strptime(f"{value}-01", "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Month format must be YYYY-MM")
    return value


def _find_inventory_by_name(db: Session, medicine_name: str) -> MedicineInventoryORM | None:
    return (
        db.query(MedicineInventoryORM)
        .filter(MedicineInventoryORM.name.ilike(medicine_name.strip()))
        .first()
    )


def _append_pdf_letterhead(elements: list, doc: SimpleDocTemplate, title: str, subtitle: str | None, styles) -> None:
    letterhead = letterhead_flowable(doc.width)
    if letterhead:
        elements.append(letterhead)
        elements.append(Spacer(1, 8))
    elements.append(Paragraph(title, styles["Title"]))
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))
    elements.append(
        Paragraph(
            f"Tanggal Cetak: {datetime.now().strftime('%d-%m-%Y %H:%M')}",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 12))


def _append_pdf_signature(elements: list, doc: SimpleDocTemplate, current_user: UserORM, styles, label: str = "Petugas UKS") -> None:
    generated_at = datetime.now()
    signer_name = current_user.full_name or "-"
    signer_nip = getattr(current_user, "nip", None) or "-"
    signer_title = getattr(current_user, "jabatan", None) or label
    qr_payload = "\n".join(
        [
            f"Nama: {signer_name}",
            f"NIP: {signer_nip}",
            f"Jabatan: {signer_title}",
            f"Tanggal cetak: {generated_at.strftime('%d/%m/%Y')}",
        ]
    )
    signature_style = styles["Normal"]
    signature_qr = qr_code_flowable(qr_payload, size=58)
    signature_qr.hAlign = "RIGHT"
    table = Table(
        [
            [
                "",
                [
                    Paragraph(f"Bekasi, {generated_at.strftime('%d/%m/%Y')}", signature_style),
                    Paragraph(signer_title, signature_style),
                    signature_qr,
                    Paragraph(signer_name, signature_style),
                    Paragraph(f"NIP. {signer_nip}", signature_style),
                ],
            ]
        ],
        colWidths=[doc.width - 190, 190],
    )
    table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    elements.append(Spacer(1, 18))
    elements.append(table)


def normalize_whatsapp_number(phone: str | None) -> str | None:
    if not phone:
        return None
    digits = "".join(ch for ch in phone if ch.isdigit())
    if not digits:
        return None
    if digits.startswith("0"):
        digits = "62" + digits[1:]
    elif digits.startswith("8"):
        digits = "62" + digits
    if len(digits) < 10:
        return None
    return digits


def send_whatsapp_message(target_phone: str | None, message: str) -> tuple[str, str]:
    token = os.getenv("FONNTE_TOKEN")
    if not token:
        return "skipped", "FONNTE_TOKEN belum diisi"

    target = normalize_whatsapp_number(target_phone)
    if not target:
        return "skipped", "Nomor wali asuh belum valid"

    try:
        response = requests.post(
            os.getenv("FONNTE_API_URL", "https://api.fonnte.com/send"),
            headers={"Authorization": token},
            data={"target": target, "message": message},
            timeout=10,
        )
        if response.ok:
            return "sent", response.text[:300]
        return "failed", f"HTTP {response.status_code}: {response.text[:300]}"
    except Exception as exc:
        return "failed", str(exc)


def build_uks_visit_whatsapp_message(patient: PatientORM, visit: UKSVisitORM) -> str:
    parent_name = patient.parent_name or "Wali Asuh / Orang Tua"
    return f"""[UKS SRMA 13 Bekasi]

Yth. {parent_name},

Siswa atas nama {patient.name} tercatat melakukan kunjungan ke UKS.

Tanggal:
{visit.visit_date}

Keluhan:
{visit.complaint}

Diagnosa:
{visit.diagnosis or "-"}

Tindakan:
{visit.treatment}

Pesan ini adalah notifikasi otomatis dari sistem UKS.

Terima kasih.
- UKS Sekolah Rakyat"""


def build_referral_whatsapp_message(patient: PatientORM, visit: UKSVisitORM) -> str:
    parent_name = patient.parent_name or "Wali Asuh / Orang Tua"
    return f"""[UKS SRMA 13 Bekasi]

Yth. {parent_name},

Siswa {patient.name} membutuhkan tindak lanjut/rujukan.

Tanggal kunjungan: {visit.visit_date}
Keluhan: {visit.complaint}
Diagnosa: {visit.diagnosis or "-"}
Tujuan rujukan: {visit.referral_to or visit.referral_place or "-"}

Mohon dilakukan pemantauan dan tindak lanjut sesuai arahan petugas UKS."""


def build_control_whatsapp_message(patient: PatientORM, visit: UKSVisitORM) -> str:
    parent_name = patient.parent_name or "Wali Asuh / Orang Tua"
    return f"""[UKS SRMA 13 Bekasi]

Yth. {parent_name},

Pengingat jadwal kontrol siswa {patient.name}.

Tanggal kontrol: {visit.control_date or "-"}
Tempat kontrol: {visit.referral_place or visit.referral_to or "-"}
Diagnosa: {visit.diagnosis or "-"}

Mohon wali asuh/orang tua memastikan jadwal kontrol terlaksana."""


def build_rest_letter_whatsapp_message(patient: PatientORM, visit: UKSVisitORM) -> str:
    parent_name = patient.parent_name or "Wali Asuh / Orang Tua"
    return f"""[UKS SRMA 13 Bekasi]

Yth. {parent_name},

Surat izin istirahat UKS untuk siswa {patient.name} telah dibuat.

Tanggal kunjungan: {visit.visit_date}
Keluhan: {visit.complaint}
Diagnosa: {visit.diagnosis or "-"}

Silakan cek surat izin dari petugas UKS."""


@router.post("/auth/register", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
def register_user(
    payload: UserCreate,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles(ROLE_ADMIN)),
) -> UserResponse:
    existing = db.query(UserORM).filter(UserORM.username == payload.username).first()
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")

    user = UserORM(
        username=payload.username,
        full_name=payload.full_name,
        role=payload.role,
        password_hash=hash_password(payload.password),
        is_active=True,
    )
    db.add(user)
    db.flush()
    write_audit_log(db, current_user, "create_user", "user", user.id, f"Created user {user.username}")
    db.commit()
    db.refresh(user)

    return _user_response(user)


@router.post("/auth/login", response_model=TokenResponse)
def login(payload: LoginRequest, request: Request, response: Response, db: Session = Depends(get_db)) -> TokenResponse:
    user = db.query(UserORM).filter(UserORM.username == payload.username).first()
    if user is None or not verify_password(payload.password, user.password_hash):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid username or password")

    if not user.is_active:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="User inactive")

    token = create_access_token(subject=str(user.id), role=user.role)
    request.session["user"] = {
        "id": user.id,
        "username": user.username,
        "full_name": user.full_name,
        "role": user.role,
    }
    response.set_cookie(
        "access_token",
        token,
        httponly=True,
        samesite="lax",
        secure=os.getenv("SESSION_COOKIE_SECURE", "false").lower() == "true",
        max_age=get_access_token_expire_seconds(),
    )
    write_audit_log(db, user, "login", "session", user.id, "User logged in")
    db.commit()
    return TokenResponse(access_token=token, expires_in=get_access_token_expire_seconds())


@router.post("/auth/logout")
def logout(
    request: Request,
    response: Response,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(get_current_user),
) -> dict:
    write_audit_log(db, current_user, "logout", "session", current_user.id, "User logged out")
    db.commit()
    request.session.clear()
    response.delete_cookie("access_token")
    return {"message": "Logged out"}


@router.post("/auth/refresh", response_model=TokenResponse)
def refresh_token(current_user: UserORM = Depends(get_current_user)) -> TokenResponse:
    token = create_access_token(subject=str(current_user.id), role=current_user.role)
    return TokenResponse(access_token=token, expires_in=get_access_token_expire_seconds())


@router.get("/auth/me", response_model=UserResponse)
def get_me(request: Request, current_user: UserORM = Depends(get_current_user)) -> UserResponse:
    request.session["user"] = {
        "id": current_user.id,
        "username": current_user.username,
        "full_name": current_user.full_name,
        "role": current_user.role,
    }
    return _user_response(current_user)


@router.post("/auth/change-password")
def change_password(
    payload: ChangePasswordRequest,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(get_current_user),
) -> dict:
    if not verify_password(payload.current_password, current_user.password_hash):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Current password is incorrect")

    current_user.password_hash = hash_password(payload.new_password)
    db.add(current_user)
    db.commit()
    return {"message": "Password updated successfully"}


@router.patch("/auth/profile", response_model=UserResponse)
def update_profile(
    payload: UserProfileUpdate,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(get_current_user),
) -> UserResponse:
    if payload.full_name is not None:
        current_user.full_name = payload.full_name
    if payload.nip is not None:
        current_user.nip = payload.nip
    if payload.jabatan is not None:
        current_user.jabatan = payload.jabatan
    if payload.signature_image is not None:
        if payload.signature_image and not payload.signature_image.startswith("data:image/"):
            raise HTTPException(status_code=400, detail="Signature image must be a PNG/JPG data URL")
        current_user.signature_image = payload.signature_image

    db.add(current_user)
    db.commit()
    db.refresh(current_user)
    return _user_response(current_user)


@router.post("/assessment", response_model=AssessmentResponse)
def assess_patient(
    payload: NursingAssessment,
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> AssessmentResponse:
    recommendations = expert_system.infer(payload)

    patient = db.get(PatientORM, payload.patient.id)
    if patient is None:
        patient = PatientORM(
            id=payload.patient.id,
            name=payload.patient.name,
            age=payload.patient.age,
            gender=payload.patient.gender,
        )
        db.add(patient)
    else:
        patient.name = payload.patient.name
        patient.age = payload.patient.age
        patient.gender = payload.patient.gender

    assessment = AssessmentORM(
        patient_id=payload.patient.id,
        complaints=payload.complaints,
        observations=payload.observations,
        vital_signs=payload.vital_signs,
    )
    db.add(assessment)
    db.flush()

    for rec in recommendations:
        db.add(
            RecommendationORM(
                assessment_id=assessment.id,
                nanda_code=rec.nanda_code,
                nanda_label=rec.nanda_label,
                confidence=rec.confidence,
                nic=rec.nic,
                noc=rec.noc,
            )
        )

    db.commit()

    return AssessmentResponse(
        patient_id=payload.patient.id,
        recommendations=recommendations,
    )

@router.post("/ai/suggest-care", response_model=AICareSuggestionResponse)
def suggest_care_with_ai(
    payload: AICareSuggestionRequest,
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> AICareSuggestionResponse:

    prompt = f"""
Anda adalah perawat UKS sekolah di Indonesia.

Buat saran klinis UKS yang praktis, spesifik, dan tidak monoton.
Gunakan istilah diagnosis keperawatan/SDKI bila sesuai, bukan diagnosis medis dokter.
Sesuaikan dengan keluhan dan hasil pemeriksaan, jangan mengulang template yang sama untuk semua kasus.

Keluhan siswa:
{payload.complaint}

Hasil pemeriksaan UKS:
{payload.examination}

Jawab dengan format persis:

Diagnosa Keperawatan:
1-2 diagnosis keperawatan yang paling relevan.

Tindakan Keperawatan:
2-4 tindakan praktis yang bisa dilakukan petugas UKS.

Catatan:
Implementasi singkat, hal yang perlu dipantau, dan kapan perlu hubungi wali/rujuk.

Tetap ringkas, aman, dan cocok untuk dokumentasi UKS.
"""

    result = ""
    ai_source = "local_fallback"
    model_used = None
    if client is not None and os.getenv("OPENROUTER_API_KEY"):
        for model_name in get_openrouter_models():
            try:
                response = client.chat.completions.create(
                    model=model_name,
                    messages=[
                        {
                            "role": "user",
                            "content": prompt,
                        }
                    ],
                    temperature=0.55,
                    max_tokens=320,
                )
                result = response.choices[0].message.content or ""
                result = result.replace("**", "")
                if result.strip():
                    ai_source = "online"
                    model_used = model_name
                    break
            except Exception:
                continue

    if not result.strip():
        result = build_local_care_suggestion(payload.complaint, payload.examination)

    diagnosis = ""
    intervention = ""
    implementation = ""

    parts = result.split("Tindakan Keperawatan:")

    if len(parts) > 1:

        diagnosis = parts[0].replace(
            "Diagnosa Keperawatan:",
            ""
        ).strip()

        tindakan_parts = parts[1].split(
            "Catatan:"
        )

        intervention = tindakan_parts[0].strip()

        if len(tindakan_parts) > 1:

            implementation = tindakan_parts[1].strip()

    else:

        diagnosis = result

    if not diagnosis:
        diagnosis = "Gangguan kenyamanan akut."
    if not intervention:
        intervention = "Observasi tanda vital, anjurkan istirahat, dan berikan cairan oral sesuai kondisi."
    if not implementation:
        implementation = "Pantau respons siswa 15-30 menit dan dokumentasikan perubahan kondisi."

    return AICareSuggestionResponse(
        diagnosis=diagnosis,
        intervention=intervention,
        implementation=implementation,
        follow_up="Evaluasi ulang sesuai kondisi pasien.",
        confidence=0.95 if ai_source == "online" else 0.75,
        source=ai_source,
        model=model_used,
    )
@router.post("/patients", response_model=PatientSummary, status_code=status.HTTP_201_CREATED)
def create_patient(
    payload: PatientCreate,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles(ROLE_ADMIN, ROLE_PERAWAT)),
) -> PatientSummary:
    existing = db.get(PatientORM, payload.id)
    if existing is not None:
        raise HTTPException(status_code=400, detail="Patient ID already exists")

    patient = PatientORM(
        id=payload.id,
        name=payload.name,
        age=payload.age,
        gender=payload.gender,
        class_name=payload.class_name,
        birth_date=payload.birth_date,
        parent_name=payload.parent_name,
        parent_phone=payload.parent_phone,
    )
    db.add(patient)
    write_audit_log(db, current_user, "create_patient", "patient", patient.id, f"Created patient {patient.name}")
    db.commit()
    db.refresh(patient)
    return PatientSummary(
        id=patient.id,
        name=patient.name,
        age=patient.age,
        gender=patient.gender,
        class_name=patient.class_name,
        birth_date=patient.birth_date,
        parent_name=patient.parent_name,
        parent_phone=patient.parent_phone,
    )


@router.get("/patients", response_model=list[PatientSummary])
def get_patients(
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> list[PatientSummary]:

    patients = (
        db.query(PatientORM)
        .order_by(PatientORM.name.asc())
        .all()
    )

    return [

        PatientSummary(
            id=p.id,
            name=p.name,
            age=p.age,
            gender=p.gender,
            class_name=p.class_name,
            birth_date=p.birth_date,
            parent_name=p.parent_name,
            parent_phone=p.parent_phone,
        )

        for p in patients

    ]


@router.get("/patients/search", response_model=list[PatientSummary])
def search_patients(
    q: str,
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> list[PatientSummary]:
    keyword = q.strip()
    if not keyword:
        raise HTTPException(status_code=400, detail="Query cannot be empty")

    like_expr = f"%{keyword}%"
    patients = (
        db.query(PatientORM)
        .filter((PatientORM.id.ilike(like_expr)) | (PatientORM.name.ilike(like_expr)))
        .order_by(PatientORM.name.asc())
        .all()
    )
    return [
        PatientSummary(
            id=p.id,
            name=p.name,
            age=p.age,
            gender=p.gender,
            class_name=p.class_name,
            birth_date=p.birth_date,
            parent_name=p.parent_name,
            parent_phone=p.parent_phone,
        )
        for p in patients
    ]


@router.get("/patients/{patient_id}", response_model=PatientSummary)
def get_patient_detail(
    patient_id: str,
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> PatientSummary:
    patient = db.get(PatientORM, patient_id)
    if patient is None:
        raise HTTPException(status_code=404, detail="Patient not found")
    return PatientSummary(
        id=patient.id,
        name=patient.name,
        age=patient.age,
        gender=patient.gender,
        class_name=patient.class_name,
        birth_date=patient.birth_date,
        parent_name=patient.parent_name,
        parent_phone=patient.parent_phone,
    )
@router.put("/patients/{patient_id}")
def update_patient(
    patient_id: str,
    payload: PatientCreate,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(
        require_roles(
            ROLE_ADMIN,
            ROLE_PERAWAT
        )
    ),
):

    patient = db.get(
        PatientORM,
        patient_id
    )

    if patient is None:

        raise HTTPException(
            status_code=404,
            detail="Patient not found"
        )

    patient.name = payload.name
    patient.age = payload.age
    patient.gender = payload.gender
    patient.class_name = payload.class_name
    patient.birth_date = payload.birth_date
    patient.parent_name = payload.parent_name
    patient.parent_phone = payload.parent_phone

    write_audit_log(db, current_user, "edit_patient", "patient", patient.id, f"Edited patient {patient.name}")
    db.commit()
    db.refresh(patient)

    return {
        "message":
        "Patient updated"
    }


@router.delete("/patients/{patient_id}")
def delete_patient(
    patient_id: str,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles(ROLE_ADMIN, ROLE_PERAWAT)),
):
    patient = db.get(PatientORM, patient_id)
    if patient is None:
        raise HTTPException(status_code=404, detail="Patient not found")

    patient_name = patient.name
    ckg_students = (
        db.query(CKGStudentORM)
        .filter(CKGStudentORM.nis == patient_id)
        .all()
    )
    for ckg_student in ckg_students:
        db.delete(ckg_student)

    db.query(RecommendationLetterORM).filter(
        RecommendationLetterORM.student_id == patient_id
    ).delete(synchronize_session=False)
    db.delete(patient)
    write_audit_log(
        db,
        current_user,
        "delete_patient",
        "patient",
        patient_id,
        f"Deleted patient {patient_name}; CKG records={len(ckg_students)}",
    )
    db.commit()
    return {
        "message": "Data siswa berhasil dihapus",
        "deleted_ckg_records": len(ckg_students),
    }

@router.post("/uks/visits", response_model=UKSVisitResponse, status_code=status.HTTP_201_CREATED)
def create_uks_visit(
    payload: UKSVisitCreate,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles(ROLE_ADMIN, ROLE_PERAWAT)),
) -> UKSVisitResponse:
    patient = db.get(PatientORM, payload.patient_id)
    if patient is None:
        raise HTTPException(status_code=404, detail="Patient not found")

    visit = UKSVisitORM(
        patient_id=payload.patient_id,
        visit_date=payload.visit_date,
        complaint=payload.complaint,
        examination=payload.examination,
        treatment=payload.treatment,
        diagnosis=payload.diagnosis,
        notes=payload.notes,
        referral_to=payload.referral_to,
        referral_status=payload.referral_status,
    )
    db.add(visit)
    db.flush()
    write_audit_log(
        db,
        current_user,
        "create_uks_visit",
        "uks_visit",
        visit.id,
        f"Created UKS visit for patient {visit.patient_id}",
    )

    whatsapp_status = "skipped"
    whatsapp_message = "Nomor wali asuh belum diisi"
    if patient and patient.parent_phone:
        whatsapp_status, whatsapp_message = send_whatsapp_message(
            patient.parent_phone,
            build_uks_visit_whatsapp_message(patient, visit),
        )
    visit.whatsapp_status = whatsapp_status
    visit.whatsapp_message = whatsapp_message
    db.add(visit)
    db.commit()
    db.refresh(visit)

    return UKSVisitResponse(
        id=visit.id,
        patient_id=visit.patient_id,
        visit_date=visit.visit_date,
        complaint=visit.complaint,
        examination=visit.examination,
        treatment=visit.treatment,
        diagnosis=visit.diagnosis,
        notes=visit.notes,
        referral_to=visit.referral_to,
        referral_status=visit.referral_status,
        whatsapp_status=whatsapp_status,
        whatsapp_message=whatsapp_message,
    )
    
@router.get("/patients/{patient_id}/visits")
def get_patient_visits(
    patient_id: str,
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
):

    visits = (
        db.query(UKSVisitORM)
        .filter(UKSVisitORM.patient_id == patient_id)
        .order_by(UKSVisitORM.visit_date.desc())
        .all()
    )

    return visits

@router.get("/uks/visits")
def get_all_uks_visits(
    month: str | None = None,
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
):
    query = db.query(UKSVisitORM)
    if month:
        month = _validate_month_yyyy_mm(month)
        query = query.filter(UKSVisitORM.visit_date.like(f"{month}%"))

    visits = (
        query
        .order_by(UKSVisitORM.id.desc())
        .all()
    )

    results = []

    for visit in visits:

        patient = db.get(PatientORM, visit.patient_id)

        results.append({
            "id": visit.id,
            "patient_id": visit.patient_id,
            "patient_name": patient.name if patient else visit.patient_id,
            "visit_date": visit.visit_date,
            "complaint": visit.complaint,
            "diagnosis": visit.diagnosis,
            "treatment": visit.treatment,
            "referral_place": visit.referral_place,
            "control_date": visit.control_date,
            "control_done": visit.control_done,
        })

    return results
@router.delete("/uks/visits/{visit_id}")
def delete_uks_visit(
    visit_id: int,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles(ROLE_ADMIN, ROLE_PERAWAT)),
):

    visit = db.get(UKSVisitORM, visit_id)

    if visit is None:
        raise HTTPException(
            status_code=404,
            detail="Visit not found"
        )

    db.delete(visit)
    write_audit_log(db, current_user, "delete_uks_visit", "uks_visit", visit_id, "Deleted UKS visit")
    db.commit()

    return {"message": "Visit deleted"}

@router.get("/uks/visits/{visit_id}", response_model=UKSVisitResponse)
def get_uks_visit_detail(
    visit_id: int,
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> UKSVisitResponse:
    visit = db.get(UKSVisitORM, visit_id)
    if visit is None:
        raise HTTPException(status_code=404, detail="UKS visit not found")

    return UKSVisitResponse(
        id=visit.id,
        patient_id=visit.patient_id,
        visit_date=visit.visit_date,
        complaint=visit.complaint,
        examination=visit.examination,
        treatment=visit.treatment,
        diagnosis=visit.diagnosis,
        notes=visit.notes,
        referral_to=visit.referral_to,
        referral_status=visit.referral_status,
    )


@router.get("/patients/{patient_id}/uks-visits", response_model=list[UKSVisitResponse])
def list_patient_uks_visits(
    patient_id: str,
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> list[UKSVisitResponse]:
    patient = db.get(PatientORM, patient_id)
    if patient is None:
        raise HTTPException(status_code=404, detail="Patient not found")

    visits = (
        db.query(UKSVisitORM)
        .filter(UKSVisitORM.patient_id == patient_id)
        .order_by(UKSVisitORM.id.desc())
        .all()
    )
    return [
        UKSVisitResponse(
            id=v.id,
            patient_id=v.patient_id,
            visit_date=v.visit_date,
            complaint=v.complaint,
            examination=v.examination,
            treatment=v.treatment,
            diagnosis=v.diagnosis,
            notes=v.notes,
            referral_to=v.referral_to,
            referral_status=v.referral_status,
        )
        for v in visits
    ]


@router.post(
    "/uks/visits/{visit_id}/medications",
    response_model=UKSMedicationResponse,
    status_code=status.HTTP_201_CREATED,
)
def add_uks_medication(
    visit_id: int,
    payload: UKSMedicationCreate,
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> UKSMedicationResponse:

    visit = db.get(UKSVisitORM, visit_id)
    if visit is None:
        raise HTTPException(status_code=404, detail="UKS visit not found")

    inventory = _find_inventory_by_name(db, payload.medicine_name)
    if inventory is None:
        raise HTTPException(status_code=404, detail="Medicine not found in inventory")

    if inventory.stock < payload.quantity:
        raise HTTPException(
            status_code=400,
            detail=f"Insufficient stock for {inventory.name}. Remaining: {inventory.stock}",
        )

    medication = UKSMedicationORM(
        visit_id=visit_id,
        medicine_name=inventory.name,
        dosage=payload.dosage,
        quantity=payload.quantity,
        notes=payload.notes,
    )

    inventory.stock -= payload.quantity

    transaction = MedicineTransactionORM(
        medicine_name=inventory.name,
        transaction_type="OUT",
        quantity=payload.quantity,
        notes=f"Kunjungan UKS #{visit_id}",
    )

    db.add(medication)
    db.add(inventory)
    db.add(transaction)

    db.commit()
    db.refresh(medication)
    db.refresh(inventory)

    return UKSMedicationResponse(
        id=medication.id,
        visit_id=medication.visit_id,
        medicine_name=medication.medicine_name,
        dosage=medication.dosage,
        quantity=medication.quantity,
        notes=medication.notes,
        remaining_stock=inventory.stock,
    )
@router.get("/uks/visits/{visit_id}/medications", response_model=list[UKSMedicationResponse])
def list_uks_medications(
    visit_id: int,
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> list[UKSMedicationResponse]:
    visit = db.get(UKSVisitORM, visit_id)
    if visit is None:
        raise HTTPException(status_code=404, detail="UKS visit not found")

    medications = (
        db.query(UKSMedicationORM)
        .filter(UKSMedicationORM.visit_id == visit_id)
        .order_by(UKSMedicationORM.id.asc())
        .all()
    )
    return [
        UKSMedicationResponse(
            id=m.id,
            visit_id=m.visit_id,
            medicine_name=m.medicine_name,
            dosage=m.dosage,
            quantity=m.quantity,
            notes=m.notes,
            remaining_stock=None,
        )
        for m in medications
    ]


@router.post(
    "/medicines",
    response_model=MedicineInventoryResponse,
    status_code=status.HTTP_201_CREATED
)
def create_medicine_inventory(
    payload: MedicineInventoryCreate,
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> MedicineInventoryResponse:

    existing = _find_inventory_by_name(
        db,
        payload.name
    )

    # tambah stok jika obat sudah ada
    if existing is not None:

        existing.stock += payload.stock
        existing.unit = payload.unit.strip()
        existing.minimum_stock = payload.minimum_stock

        transaction = MedicineTransactionORM(
            medicine_name=existing.name,
            transaction_type="IN",
            quantity=payload.stock,
            notes="Penambahan stok"
        )

        db.add(transaction)

        db.commit()
        db.refresh(existing)

        return MedicineInventoryResponse(
            id=existing.id,
            name=existing.name,
            unit=existing.unit,
            stock=existing.stock,
            minimum_stock=existing.minimum_stock,
            is_low_stock=(
                existing.stock <= existing.minimum_stock
            ),
        )

    # buat obat baru
    med = MedicineInventoryORM(
        name=payload.name.strip(),
        unit=payload.unit.strip(),
        stock=payload.stock,
        minimum_stock=payload.minimum_stock,
    )

    db.add(med)

    transaction = MedicineTransactionORM(
        medicine_name=payload.name.strip(),
        transaction_type="IN",
        quantity=payload.stock,
        notes="Stok awal"
    )

    db.add(transaction)

    db.commit()
    db.refresh(med)

    return MedicineInventoryResponse(
        id=med.id,
        name=med.name,
        unit=med.unit,
        stock=med.stock,
        minimum_stock=med.minimum_stock,
        is_low_stock=(
            med.stock <= med.minimum_stock
        ),
    )

@router.get("/medicines", response_model=list[MedicineInventoryResponse])
def list_medicines_inventory(
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> list[MedicineInventoryResponse]:
    meds = db.query(MedicineInventoryORM).order_by(MedicineInventoryORM.name.asc()).all()
    return [
        MedicineInventoryResponse(
            id=m.id,
            name=m.name,
            unit=m.unit,
            stock=m.stock,
            minimum_stock=m.minimum_stock,
            is_low_stock=m.stock <= m.minimum_stock,
        )
        for m in meds
    ]


@router.patch("/medicines/{medicine_id}", response_model=MedicineInventoryResponse)
def update_medicine_inventory(
    medicine_id: int,
    payload: MedicineInventoryUpdate,
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> MedicineInventoryResponse:
    med = db.get(MedicineInventoryORM, medicine_id)
    if med is None:
        raise HTTPException(status_code=404, detail="Medicine not found")

    if payload.unit is not None:
        med.unit = payload.unit.strip()
    if payload.stock is not None:
        med.stock = payload.stock
    if payload.minimum_stock is not None:
        med.minimum_stock = payload.minimum_stock

    db.add(med)
    db.commit()
    db.refresh(med)
    return MedicineInventoryResponse(
        id=med.id,
        name=med.name,
        unit=med.unit,
        stock=med.stock,
        minimum_stock=med.minimum_stock,
        is_low_stock=med.stock <= med.minimum_stock,
    )


@router.post("/medicines/{medicine_id}/adjust", response_model=MedicineInventoryResponse)
def adjust_medicine_stock(
    medicine_id: int,
    payload: MedicineStockAdjustment,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles("admin", "perawat")),
) -> MedicineInventoryResponse:
    med = db.get(MedicineInventoryORM, medicine_id)
    if med is None:
        raise HTTPException(status_code=404, detail="Medicine not found")

    before_stock = med.stock
    transaction_type = payload.adjustment_type
    transaction_quantity = payload.quantity

    if payload.adjustment_type == "IN":
        if payload.quantity < 1:
            raise HTTPException(status_code=400, detail="Quantity must be greater than 0")
        med.stock += payload.quantity
    elif payload.adjustment_type == "OUT":
        if payload.quantity < 1:
            raise HTTPException(status_code=400, detail="Quantity must be greater than 0")
        if med.stock < payload.quantity:
            raise HTTPException(status_code=400, detail=f"Insufficient stock for {med.name}. Remaining: {med.stock}")
        med.stock -= payload.quantity
    else:
        if payload.quantity == before_stock:
            transaction_type = "IN"
            transaction_quantity = 0
        elif payload.quantity > before_stock:
            transaction_type = "IN"
            transaction_quantity = payload.quantity - before_stock
        else:
            transaction_type = "OUT"
            transaction_quantity = before_stock - payload.quantity
        med.stock = payload.quantity

    transaction = MedicineTransactionORM(
        medicine_name=med.name,
        transaction_type=transaction_type,
        quantity=transaction_quantity,
        notes=payload.notes or f"Koreksi stok dari {before_stock} ke {med.stock}",
    )
    db.add(med)
    db.add(transaction)
    write_audit_log(
        db,
        current_user,
        "adjust_medicine_stock",
        "medicine",
        med.id,
        f"{med.name}: {before_stock} -> {med.stock}",
    )
    db.commit()
    db.refresh(med)

    return MedicineInventoryResponse(
        id=med.id,
        name=med.name,
        unit=med.unit,
        stock=med.stock,
        minimum_stock=med.minimum_stock,
        is_low_stock=med.stock <= med.minimum_stock,
    )


@router.get("/reports/medicine-mutation")
def get_medicine_mutation_report(
    month: int,
    year: int,
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> list[dict]:
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="Month must be 1-12")
    start_dt = datetime(year, month, 1)
    end_dt = datetime(year + 1, 1, 1) if month == 12 else datetime(year, month + 1, 1)

    transactions = (
        db.query(MedicineTransactionORM)
        .filter(MedicineTransactionORM.transaction_date >= start_dt)
        .filter(MedicineTransactionORM.transaction_date < end_dt)
        .order_by(MedicineTransactionORM.medicine_name.asc())
        .all()
    )
    summary: dict[str, dict] = {}
    for trx in transactions:
        item = summary.setdefault(
            trx.medicine_name,
            {"medicine_name": trx.medicine_name, "in_qty": 0, "out_qty": 0, "current_stock": 0},
        )
        if trx.transaction_type == "IN":
            item["in_qty"] += trx.quantity
        else:
            item["out_qty"] += trx.quantity

    stocks = {
        med.name: med.stock
        for med in db.query(MedicineInventoryORM).order_by(MedicineInventoryORM.name.asc()).all()
    }
    for name, stock in stocks.items():
        item = summary.setdefault(
            name,
            {"medicine_name": name, "in_qty": 0, "out_qty": 0, "current_stock": stock},
        )
        item["current_stock"] = stock

    return list(summary.values())


@router.get("/reports/medicines/pdf")
def get_medicines_report_pdf(
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(
        require_roles("admin", "perawat")
    ),
) -> StreamingResponse:
    medicines = db.query(MedicineInventoryORM).order_by(MedicineInventoryORM.name.asc()).all()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()

    elements = []
    _append_pdf_letterhead(
        elements,
        doc,
        "LAPORAN STOK OBAT UKS",
        f"Total item: {len(medicines)}",
        styles,
    )

    data = [["No", "Nama Obat", "Stok", "Satuan", "Stok Minimum", "Status"]]
    for idx, med in enumerate(medicines, start=1):
        status_text = "Perlu Restok" if med.stock <= med.minimum_stock else "Aman"
        data.append(
            [
                str(idx),
                med.name,
                str(med.stock),
                med.unit,
                str(med.minimum_stock),
                status_text,
            ]
        )

    table = Table(data, repeatRows=1, colWidths=[28, 220, 50, 70, 80, 90])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("ALIGN", (1, 1), (1, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#94a3b8")),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements.append(table)
    _append_pdf_signature(elements, doc, current_user, styles)
    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="laporan_stok_obat_uks.pdf"'},
    )
@router.get("/reports/medicine-mutation/pdf")
def get_medicine_mutation_pdf(
    month: int,
    year: int,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(
    require_roles("admin", "perawat")
    ),
) -> StreamingResponse:


    start_dt = datetime(year, month, 1)

    end_dt = (
        datetime(year + 1, 1, 1)
        if month == 12
        else datetime(year, month + 1, 1)
)

    transactions = (
        db.query(MedicineTransactionORM)
        .filter(MedicineTransactionORM.transaction_date >= start_dt)
        .filter(MedicineTransactionORM.transaction_date < end_dt)
        .order_by(MedicineTransactionORM.transaction_date.asc())
        .all()
)

    mutation_rows = []

    for trx in transactions:

        stock_item = (
            db.query(MedicineInventoryORM)
            .filter(
                MedicineInventoryORM.name == trx.medicine_name
            )
            .first()
        )

        mutation_rows.append(
            {
                "date": trx.transaction_date.strftime("%d-%m-%Y"),
                "medicine": trx.medicine_name,
                "type": "Masuk" if trx.transaction_type == "IN" else "Keluar",
                "qty": trx.quantity,
                "stock": stock_item.stock if stock_item else 0,
            }
        )

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()

    elements = []
    _append_pdf_letterhead(
        elements,
        doc,
        "LAPORAN MUTASI OBAT UKS",
        f"Periode: {month:02d}/{year}",
        styles,
    )

    data = [
        [
            "No",
            "Tanggal",
            "Nama Obat",
            "Jenis",
            "Jumlah",
            "Stok Saat Ini",
        ]
    ]

    for idx, row in enumerate(mutation_rows, start=1):
        data.append(
            [
                str(idx),
                row["date"],
                row["medicine"],
                row["type"],
                str(row["qty"]),
                str(row["stock"]),
            ]
        )

    table = Table(
        data,
        repeatRows=1,
        colWidths=[30, 70, 210, 70, 60, 80]
    )

    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
            ]
        )
    )

    elements.append(table)
    _append_pdf_signature(elements, doc, current_user, styles)

    doc.build(elements)

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
            f'attachment; filename="laporan_mutasi_obat_{year}_{month:02d}.pdf"'
        },
    )
@router.patch("/uks/visits/{visit_id}/referral", response_model=UKSVisitResponse)
def update_uks_referral(
    visit_id: int,
    payload: UKSReferralUpdate,
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> UKSVisitResponse:
    visit = db.get(UKSVisitORM, visit_id)
    if visit is None:
        raise HTTPException(status_code=404, detail="UKS visit not found")

    visit.referral_to = payload.referral_to
    visit.referral_status = payload.referral_status
    db.add(visit)
    db.commit()
    db.refresh(visit)
    patient = db.get(PatientORM, visit.patient_id)

    if patient and patient.parent_phone:
        send_whatsapp_message(
            patient.parent_phone,
            build_referral_whatsapp_message(patient, visit),
        )

    return UKSVisitResponse(
        id=visit.id,
        patient_id=visit.patient_id,
        visit_date=str(visit.visit_date),
        complaint=visit.complaint,
        examination=visit.examination,
        treatment=visit.treatment,
        diagnosis=visit.diagnosis,
        notes=visit.notes,
        referral_to=visit.referral_to,
        referral_status=visit.referral_status,
    )

    return UKSVisitResponse(
        id=visit.id,
        patient_id=visit.patient_id,
        visit_date=visit.visit_date,
        complaint=visit.complaint,
        examination=visit.examination,
        treatment=visit.treatment,
        diagnosis=visit.diagnosis,
        notes=visit.notes,
        referral_to=visit.referral_to,
        referral_status=visit.referral_status,
    )


@router.get("/reports/uks/daily", response_model=UKSDailyReportResponse)
def get_uks_daily_report(
    date: str,
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> UKSDailyReportResponse:
    date = _validate_date_yyyy_mm_dd(date)
    visits = db.query(UKSVisitORM).filter(UKSVisitORM.visit_date == date).all()
    total_referrals = sum(1 for visit in visits if visit.referral_status == "dirujuk")
    return UKSDailyReportResponse(
        date=date,
        total_visits=len(visits),
        total_referrals=total_referrals,
        top_complaints=_build_top_complaints(visits),
    )


@router.get("/reports/uks/daily/excel")
def get_uks_daily_report_excel(
    date: str,
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> StreamingResponse:
    date = _validate_date_yyyy_mm_dd(date)
    visits = (
        db.query(UKSVisitORM)
        .filter(UKSVisitORM.visit_date == date)
        .order_by(UKSVisitORM.id.asc())
        .all()
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Laporan Harian"
    headers = [
        "No",
        "TANGGAL / BULAN",
        "NAMA",
        "USIA",
        "KELAS",
        "KELUHAN",
        "DIAGNOSA",
        "HASIL PEMERIKSAAN SINGKAT",
        "IMPLEMENTASI DAN RENCANA TINDAK LANJUT",
    ]
    sheet.append(headers)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    for col, _ in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for idx, visit in enumerate(visits, start=1):
        patient = db.get(PatientORM, visit.patient_id)
        try:
            formatted_date = datetime.strptime(visit.visit_date, "%Y-%m-%d").strftime("%d/%m/%Y")
        except ValueError:
            formatted_date = visit.visit_date

        diagnosis = visit.diagnosis or "-"
        hasil_pemeriksaan_singkat = visit.examination
        implementasi_dan_tindak_lanjut = (
            f"{visit.treatment}. {visit.notes}" if visit.notes else visit.treatment
        )

        sheet.append(
            [
                idx,
                formatted_date,
                patient.name if patient else visit.patient_id,
                patient.age if patient else "",
                patient.class_name if patient and patient.class_name else "",
                visit.complaint,
                diagnosis,
                hasil_pemeriksaan_singkat,
                implementasi_dan_tindak_lanjut,
            ]
        )

    widths = {
        "A": 6,
        "B": 16,
        "C": 24,
        "D": 10,
        "E": 12,
        "F": 22,
        "G": 18,
        "H": 58,
        "I": 34,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=9):
        for cell in row:
            cell.border = border
            if cell.column in (1, 2, 4, 5):
                cell.alignment = center
            else:
                cell.alignment = wrap

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"laporan_harian_uks_{date}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/reports/uks/monthly", response_model=UKSMonthlyReportResponse)
def get_uks_monthly_report(
    month: str,
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> UKSMonthlyReportResponse:
    month = _validate_month_yyyy_mm(month)
    month_prefix = f"{month}%"
    visits = db.query(UKSVisitORM).filter(UKSVisitORM.visit_date.like(month_prefix)).all()
    total_referrals = sum(1 for visit in visits if visit.referral_status == "dirujuk")
    return UKSMonthlyReportResponse(
        month=month,
        total_visits=len(visits),
        total_referrals=total_referrals,
        top_complaints=_build_top_complaints(visits),
    )


def _visit_report_bounds(
    period: str,
    report_date: str | None,
    start_date: str | None,
    end_date: str | None,
    month: str | None,
) -> tuple[str, str, str]:
    if period == "daily":
        if not report_date:
            raise HTTPException(status_code=400, detail="date is required for daily report")
        safe_date = _validate_date_yyyy_mm_dd(report_date)
        return safe_date, safe_date, f"Harian {safe_date}"
    if period == "weekly":
        if not start_date or not end_date:
            raise HTTPException(status_code=400, detail="start_date and end_date are required for weekly report")
        safe_start = _validate_date_yyyy_mm_dd(start_date)
        safe_end = _validate_date_yyyy_mm_dd(end_date)
        if safe_start > safe_end:
            raise HTTPException(status_code=400, detail="start_date cannot be after end_date")
        return safe_start, safe_end, f"Mingguan {safe_start} s/d {safe_end}"
    if period == "monthly":
        if not month:
            raise HTTPException(status_code=400, detail="month is required for monthly report")
        safe_month = _validate_month_yyyy_mm(month)
        start_dt = datetime.strptime(f"{safe_month}-01", "%Y-%m-%d")
        if start_dt.month == 12:
            end_dt = datetime(start_dt.year + 1, 1, 1) - timedelta(days=1)
        else:
            end_dt = datetime(start_dt.year, start_dt.month + 1, 1) - timedelta(days=1)
        return start_dt.strftime("%Y-%m-%d"), end_dt.strftime("%Y-%m-%d"), f"Bulanan {safe_month}"
    raise HTTPException(status_code=400, detail="period must be daily, weekly, or monthly")


def _visit_report_rows(db: Session, start: str, end: str) -> list[dict]:
    visits = (
        db.query(UKSVisitORM)
        .filter(UKSVisitORM.visit_date >= start)
        .filter(UKSVisitORM.visit_date <= end)
        .order_by(UKSVisitORM.visit_date.asc(), UKSVisitORM.id.asc())
        .all()
    )
    patient_ids = [visit.patient_id for visit in visits if visit.patient_id]
    patients = {
        patient.id: patient
        for patient in db.query(PatientORM).filter(PatientORM.id.in_(patient_ids)).all()
    } if patient_ids else {}
    rows = []
    for visit in visits:
        patient = patients.get(visit.patient_id)
        rows.append(
            {
                "tanggal": visit.visit_date,
                "nama_siswa": patient.name if patient else visit.patient_id,
                "kelas": patient.class_name if patient and patient.class_name else "-",
                "keluhan": visit.complaint or "-",
                "diagnosa": visit.diagnosis or "-",
                "tindakan": visit.treatment or "-",
                "petugas": "-",
            }
        )
    return rows


def _visit_report_payload(
    db: Session,
    period: str,
    report_date: str | None,
    start_date: str | None,
    end_date: str | None,
    month: str | None,
) -> dict:
    start, end, label = _visit_report_bounds(period, report_date, start_date, end_date, month)
    rows = _visit_report_rows(db, start, end)
    return {
        "period": period,
        "label": label,
        "start_date": start,
        "end_date": end,
        "total": len(rows),
        "rows": rows,
    }


@router.get("/reports/uks/visits")
def get_uks_visit_report(
    period: str = Query(..., pattern="^(daily|weekly|monthly)$"),
    date: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    month: str | None = None,
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> dict:
    return _visit_report_payload(db, period, date, start_date, end_date, month)


@router.get("/reports/uks/visits/pdf")
def get_uks_visit_report_pdf(
    period: str = Query(..., pattern="^(daily|weekly|monthly)$"),
    date: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    month: str | None = None,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles("admin", "perawat")),
) -> StreamingResponse:
    payload = _visit_report_payload(db, period, date, start_date, end_date, month)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=28, rightMargin=28, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    body = styles["BodyText"]
    body.fontSize = 7
    body.leading = 8.5
    title = styles["Title"]
    title.fontSize = 14
    elements = []
    _append_pdf_letterhead(
        elements,
        doc,
        "LAPORAN KUNJUNGAN UKS",
        escape(payload["label"]),
        styles,
    )
    headers = ["Tanggal", "Nama Siswa", "Kelas", "Keluhan", "Diagnosa", "Tindakan", "Petugas"]
    table_rows = [[Paragraph(escape(header), body) for header in headers]]
    for row in payload["rows"]:
        table_rows.append([Paragraph(escape(str(row[key])), body) for key in ["tanggal", "nama_siswa", "kelas", "keluhan", "diagnosa", "tindakan", "petugas"]])
    if len(table_rows) == 1:
        table_rows.append([Paragraph("Tidak ada data", body), "", "", "", "", "", ""])
    table = Table(table_rows, repeatRows=1, colWidths=[58, 130, 45, 115, 130, 235, 60])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#9ca3af")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    elements.append(table)
    _append_pdf_signature(elements, doc, current_user, styles)
    doc.build(elements)
    buffer.seek(0)
    filename = f"laporan_kunjungan_uks_{period}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/reports/uks/visits/excel")
def get_uks_visit_report_excel(
    period: str = Query(..., pattern="^(daily|weekly|monthly)$"),
    date: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    month: str | None = None,
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> StreamingResponse:
    payload = _visit_report_payload(db, period, date, start_date, end_date, month)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Laporan Kunjungan"
    headers = ["Tanggal", "Nama Siswa", "Kelas", "Keluhan", "Diagnosa", "Tindakan", "Petugas"]
    sheet.append(headers)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    for col, _ in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for row in payload["rows"]:
        sheet.append([row[key] for key in ["tanggal", "nama_siswa", "kelas", "keluhan", "diagnosa", "tindakan", "petugas"]])

    widths = {"A": 14, "B": 28, "C": 12, "D": 28, "E": 30, "F": 48, "G": 20}
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = border
            cell.alignment = center if cell.column in (1, 3, 7) else wrap

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"laporan_kunjungan_uks_{period}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/patients/{patient_id}/assessments", response_model=PatientAssessmentsResponse)
def get_patient_assessments(
    patient_id: str,
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> PatientAssessmentsResponse:
    patient = db.get(PatientORM, patient_id)
    if patient is None:
        raise HTTPException(status_code=404, detail="Patient not found")

    assessments = (
        db.query(AssessmentORM)
        .filter(AssessmentORM.patient_id == patient_id)
        .order_by(AssessmentORM.id.desc())
        .all()
    )

    payload = []
    for a in assessments:
        recs = (
            db.query(RecommendationORM)
            .filter(RecommendationORM.assessment_id == a.id)
            .order_by(RecommendationORM.id.asc())
            .all()
        )
        payload.append(
            AssessmentSummary(
                id=a.id,
                patient_id=a.patient_id,
                complaints=a.complaints,
                observations=a.observations,
                vital_signs=a.vital_signs,
                recommendations=[
                    {
                        "nanda_code": r.nanda_code,
                        "nanda_label": r.nanda_label,
                        "confidence": r.confidence,
                        "nic": r.nic,
                        "noc": r.noc,
                    }
                    for r in recs
                ],
            )
        )

    return PatientAssessmentsResponse(
        patient=PatientSummary(
            id=patient.id,
            name=patient.name,
            age=patient.age,
            gender=patient.gender,
            class_name=patient.class_name,
        ),
        assessments=payload,
    )


@router.get("/assessments/{assessment_id}", response_model=AssessmentSummary)
def get_assessment_detail(
    assessment_id: int,
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles("admin", "perawat")),
) -> AssessmentSummary:
    assessment = db.get(AssessmentORM, assessment_id)
    if assessment is None:
        raise HTTPException(status_code=404, detail="Assessment not found")

    recs = (
        db.query(RecommendationORM)
        .filter(RecommendationORM.assessment_id == assessment.id)
        .order_by(RecommendationORM.id.asc())
        .all()
    )

    return AssessmentSummary(
        id=assessment.id,
        patient_id=assessment.patient_id,
        complaints=assessment.complaints,
        observations=assessment.observations,
        vital_signs=assessment.vital_signs,
        recommendations=[
            {
                "nanda_code": r.nanda_code,
                "nanda_label": r.nanda_label,
                "confidence": r.confidence,
                "nic": r.nic,
                "noc": r.noc,
            }
            for r in recs
        ],
    )
@router.put("/uks/visits/{visit_id}")
def update_uks_visit(
    visit_id: int,
    payload: UKSVisitCreate,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles(ROLE_ADMIN, ROLE_PERAWAT)),
):

    visit = db.get(UKSVisitORM, visit_id)

    if visit is None:
        raise HTTPException(
            status_code=404,
            detail="Visit not found"
        )

    if payload.patient_id is not None:
        visit.patient_id = payload.patient_id

    if payload.visit_date is not None:
        visit.visit_date = payload.visit_date

    if payload.complaint is not None:
        visit.complaint = payload.complaint

    if payload.examination is not None:
        visit.examination = payload.examination

    if payload.treatment is not None:
        visit.treatment = payload.treatment

    if payload.diagnosis is not None:
        visit.diagnosis = payload.diagnosis

    if payload.notes is not None:
        visit.notes = payload.notes

    if payload.referral_place is not None:
        visit.referral_place = payload.referral_place

    if payload.control_date is not None:
        visit.control_date = payload.control_date

    if payload.control_done is not None:
        visit.control_done = payload.control_done

    write_audit_log(db, current_user, "edit_uks_visit", "uks_visit", visit.id, f"Edited UKS visit for patient {visit.patient_id}")
    db.commit()
    db.refresh(visit)
    if payload.control_date is not None or payload.referral_place is not None:
        patient = db.get(PatientORM, visit.patient_id)
        if patient and patient.parent_phone:
            send_whatsapp_message(patient.parent_phone, build_control_whatsapp_message(patient, visit))

    return {
        "message": "Visit updated"
    }
from sqlalchemy import func
from datetime import date


@router.get("/dashboard/stats")
def dashboard_stats(
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles(ROLE_ADMIN, ROLE_PERAWAT)),
):

    total_students = db.query(PatientORM).count()

    today_visits = db.query(UKSVisitORM).filter(
        UKSVisitORM.visit_date == date.today().isoformat()
    ).count()

    top_case = (
        db.query(
            UKSVisitORM.diagnosis,
            func.count(UKSVisitORM.diagnosis).label("total")
        )
        .group_by(UKSVisitORM.diagnosis)
        .order_by(func.count(UKSVisitORM.diagnosis).desc())
        .first()
    )

    return {
        "total_students": total_students,
        "today_visits": today_visits,
        "top_case": top_case[0] if top_case else "-",
        "active_reports": today_visits
    }
@router.get("/users")
def list_users(
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles(ROLE_ADMIN))
):

    users = (
        db.query(UserORM)
        .order_by(UserORM.full_name.asc())
        .all()
    )

    return [
        {
            "id": user.id,
            "username": user.username,
            "full_name": user.full_name,
            "role": user.role,
            "is_active": user.is_active,
            "created_at": user.created_at,
            "updated_at": user.updated_at,
        }
        for user in users
    ]


@router.post("/users", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
def create_user_from_admin(
    payload: UserCreate,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles(ROLE_ADMIN)),
) -> UserResponse:
    existing = db.query(UserORM).filter(UserORM.username == payload.username).first()
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")

    user = UserORM(
        username=payload.username,
        full_name=payload.full_name,
        role=payload.role,
        password_hash=hash_password(payload.password),
        is_active=True,
    )
    db.add(user)
    db.flush()
    write_audit_log(db, current_user, "create_user", "user", user.id, f"Created user {user.username}")
    db.commit()
    db.refresh(user)

    return _user_response(user)


@router.patch("/users/{user_id}", response_model=UserResponse)
def update_user_from_admin(
    user_id: int,
    payload: UserUpdate,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles(ROLE_ADMIN)),
) -> UserResponse:
    user = db.get(UserORM, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")

    if payload.username is not None and payload.username != user.username:
        existing = db.query(UserORM).filter(UserORM.username == payload.username).first()
        if existing is not None:
            raise HTTPException(status_code=400, detail="Username already exists")
        user.username = payload.username

    if payload.full_name is not None:
        user.full_name = payload.full_name

    if payload.role is not None:
        if user.role == ROLE_ADMIN and payload.role != ROLE_ADMIN and user.is_active and _active_admin_count(db) <= 1:
            raise HTTPException(status_code=400, detail="Cannot remove the last active admin")
        user.role = payload.role

    if payload.is_active is not None:
        if user.id == current_user.id and payload.is_active is False:
            raise HTTPException(status_code=400, detail="Cannot deactivate your own account")
        if user.role == ROLE_ADMIN and user.is_active and payload.is_active is False and _active_admin_count(db) <= 1:
            raise HTTPException(status_code=400, detail="Cannot deactivate the last active admin")
        user.is_active = payload.is_active

    write_audit_log(db, current_user, "edit_user", "user", user.id, f"Edited user {user.username}")
    db.add(user)
    db.commit()
    db.refresh(user)
    return _user_response(user)


@router.post("/users/{user_id}/reset-password")
def reset_user_password(
    user_id: int,
    payload: PasswordResetRequest,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles(ROLE_ADMIN)),
) -> dict:
    user = db.get(UserORM, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")

    user.password_hash = hash_password(payload.new_password)
    write_audit_log(db, current_user, "reset_password", "user", user.id, f"Reset password for {user.username}")
    db.add(user)
    db.commit()
    return {"message": "Password reset successfully"}


@router.post("/users/{user_id}/activate", response_model=UserResponse)
def activate_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles(ROLE_ADMIN)),
) -> UserResponse:
    user = db.get(UserORM, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")
    user.is_active = True
    write_audit_log(db, current_user, "activate_user", "user", user.id, f"Activated user {user.username}")
    db.add(user)
    db.commit()
    db.refresh(user)
    return _user_response(user)


@router.post("/users/{user_id}/deactivate", response_model=UserResponse)
def deactivate_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles(ROLE_ADMIN)),
) -> UserResponse:
    user = db.get(UserORM, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")
    if user.id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot deactivate your own account")
    if user.role == ROLE_ADMIN and user.is_active and _active_admin_count(db) <= 1:
        raise HTTPException(status_code=400, detail="Cannot deactivate the last active admin")
    user.is_active = False
    write_audit_log(db, current_user, "deactivate_user", "user", user.id, f"Deactivated user {user.username}")
    db.add(user)
    db.commit()
    db.refresh(user)
    return _user_response(user)


@router.delete("/users/{user_id}")
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles(ROLE_ADMIN)),
) -> dict:
    user = db.get(UserORM, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")
    if user.id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    if user.role == ROLE_ADMIN and user.is_active and _active_admin_count(db) <= 1:
        raise HTTPException(status_code=400, detail="Cannot delete the last active admin")
    username = user.username
    db.delete(user)
    write_audit_log(db, current_user, "delete_user", "user", user_id, f"Deleted user {username}")
    db.commit()
    return {"message": "User deleted"}


@router.get("/audit-logs", response_model=AuditLogListResponse)
def list_audit_logs(
    user: str | None = None,
    action: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    search: str | None = None,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles(ROLE_ADMIN)),
) -> AuditLogListResponse:
    query = db.query(AuditLogORM)

    if user:
        like_user = f"%{user.strip()}%"
        query = query.filter(AuditLogORM.username.ilike(like_user))
    if action:
        query = query.filter(AuditLogORM.action == action.strip())
    if date_from:
        query = query.filter(AuditLogORM.timestamp >= datetime.strptime(date_from, "%Y-%m-%d"))
    if date_to:
        query = query.filter(AuditLogORM.timestamp < datetime.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59))
    if search:
        like_search = f"%{search.strip()}%"
        query = query.filter(
            (AuditLogORM.username.ilike(like_search))
            | (AuditLogORM.action.ilike(like_search))
            | (AuditLogORM.entity_type.ilike(like_search))
            | (AuditLogORM.entity_id.ilike(like_search))
            | (AuditLogORM.details.ilike(like_search))
        )

    total = query.count()
    logs = (
        query.order_by(AuditLogORM.timestamp.desc(), AuditLogORM.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return AuditLogListResponse(
        items=[
            AuditLogResponse(
                id=log.id,
                user_id=log.user_id,
                username=log.username,
                action=log.action,
                entity_type=log.entity_type,
                entity_id=log.entity_id,
                details=log.details,
                timestamp=log.timestamp,
            )
            for log in logs
        ],
        total=total,
        page=page,
        page_size=page_size,
    )


@router.get("/dashboard/advanced-stats")
def dashboard_advanced_stats(
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles(ROLE_ADMIN, ROLE_PERAWAT)),
):
    today = date.today().isoformat()
    month_prefix = today[:7]

    visits = db.query(UKSVisitORM).all()
    monthly_visits = [v for v in visits if str(v.visit_date).startswith(month_prefix)]
    today_visits = [v for v in visits if str(v.visit_date) == today]
    low_stock = (
        db.query(MedicineInventoryORM)
        .filter(MedicineInventoryORM.stock <= MedicineInventoryORM.minimum_stock)
        .order_by(MedicineInventoryORM.name.asc())
        .limit(10)
        .all()
    )
    pending_controls = [
        v for v in visits
        if getattr(v, "control_date", None) and not getattr(v, "control_done", False)
    ]
    student_counter: dict[str, dict] = {}
    for visit in monthly_visits:
        patient = db.get(PatientORM, visit.patient_id)
        key = visit.patient_id
        if key not in student_counter:
            student_counter[key] = {
                "student_id": key,
                "name": patient.name if patient else key,
                "class_name": patient.class_name if patient else "-",
                "total": 0,
            }
        student_counter[key]["total"] += 1

    return {
        "today_visits": len(today_visits),
        "monthly_visits": len(monthly_visits),
        "top_monthly_students": sorted(student_counter.values(), key=lambda x: x["total"], reverse=True)[:10],
        "low_stock": [
            {
                "id": med.id,
                "name": med.name,
                "stock": med.stock,
                "unit": med.unit,
                "minimum_stock": med.minimum_stock,
            }
            for med in low_stock
        ],
        "pending_controls": [
            {
                "visit_id": visit.id,
                "patient_id": visit.patient_id,
                "patient_name": db.get(PatientORM, visit.patient_id).name if db.get(PatientORM, visit.patient_id) else visit.patient_id,
                "control_date": visit.control_date,
                "referral_place": visit.referral_place or visit.referral_to,
            }
            for visit in sorted(pending_controls, key=lambda v: str(v.control_date))[:10]
        ],
        "whatsapp": {
            "configured": bool(os.getenv("FONNTE_TOKEN")),
            "visits_with_parent_phone": sum(1 for v in today_visits if (db.get(PatientORM, v.patient_id) and db.get(PatientORM, v.patient_id).parent_phone)),
            "visits_without_parent_phone": sum(1 for v in today_visits if not (db.get(PatientORM, v.patient_id) and db.get(PatientORM, v.patient_id).parent_phone)),
        },
    }


@router.get("/whatsapp/logs")
def list_whatsapp_logs(
    status_filter: str | None = Query(default=None, alias="status"),
    limit: int = Query(default=20, ge=1, le=100),
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles(ROLE_ADMIN, ROLE_PERAWAT)),
) -> list[dict]:
    query = db.query(UKSVisitORM).order_by(UKSVisitORM.id.desc())
    if status_filter:
        query = query.filter(UKSVisitORM.whatsapp_status == status_filter)
    visits = query.limit(limit).all()
    rows = []
    for visit in visits:
        patient = db.get(PatientORM, visit.patient_id)
        rows.append(
            {
                "visit_id": visit.id,
                "visit_date": visit.visit_date,
                "patient_id": visit.patient_id,
                "patient_name": patient.name if patient else visit.patient_id,
                "parent_phone": patient.parent_phone if patient else None,
                "status": visit.whatsapp_status or "skipped",
                "message": visit.whatsapp_message or "-",
            }
        )
    return rows


@router.post("/whatsapp/visits/{visit_id}/resend")
def resend_visit_whatsapp(
    visit_id: int,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles(ROLE_ADMIN, ROLE_PERAWAT)),
) -> dict:
    visit = db.get(UKSVisitORM, visit_id)
    if visit is None:
        raise HTTPException(status_code=404, detail="UKS visit not found")
    patient = db.get(PatientORM, visit.patient_id)
    if patient is None:
        raise HTTPException(status_code=404, detail="Patient not found")

    status_text, detail = send_whatsapp_message(
        patient.parent_phone,
        build_uks_visit_whatsapp_message(patient, visit),
    )
    visit.whatsapp_status = status_text
    visit.whatsapp_message = detail
    db.add(visit)
    write_audit_log(db, current_user, "resend_whatsapp_visit", "uks_visit", visit.id, f"{status_text}: {detail}")
    db.commit()
    return {"whatsapp_status": status_text, "whatsapp_message": detail}


@router.get("/system/health-check")
def system_health_check(
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles(ROLE_ADMIN)),
):
    checks = []
    try:
        db.query(UserORM).count()
        checks.append({"name": "Database", "status": "ok", "detail": "Koneksi database aktif"})
    except Exception as exc:
        checks.append({"name": "Database", "status": "error", "detail": str(exc)})

    checks.append({
        "name": "WhatsApp",
        "status": "ok" if os.getenv("FONNTE_TOKEN") else "warning",
        "detail": "FONNTE_TOKEN terisi" if os.getenv("FONNTE_TOKEN") else "FONNTE_TOKEN belum diisi",
    })
    letterhead = Path("static/img/kop-surat-sekolah-rakyat.png")
    checks.append({
        "name": "Kop Surat PDF",
        "status": "ok" if letterhead.exists() else "warning",
        "detail": str(letterhead),
    })
    static_dir = Path("static")
    checks.append({
        "name": "Static Folder",
        "status": "ok" if static_dir.exists() else "error",
        "detail": str(static_dir.resolve()) if static_dir.exists() else "Folder static tidak ditemukan",
    })
    return {"status": "ok" if all(c["status"] != "error" for c in checks) else "error", "checks": checks}


@router.get("/admin/backup")
def download_backup(
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles(ROLE_ADMIN)),
):
    payload = {
        "generated_at": datetime.utcnow().isoformat(),
        "patients": [
            {
                "id": p.id,
                "name": p.name,
                "age": p.age,
                "gender": p.gender,
                "class_name": p.class_name,
                "birth_date": p.birth_date,
                "parent_name": p.parent_name,
                "parent_phone": p.parent_phone,
            }
            for p in db.query(PatientORM).all()
        ],
        "visits": [
            {
                "id": v.id,
                "patient_id": v.patient_id,
                "visit_date": v.visit_date,
                "complaint": v.complaint,
                "examination": v.examination,
                "treatment": v.treatment,
                "diagnosis": v.diagnosis,
                "notes": v.notes,
                "referral_to": v.referral_to,
                "referral_status": v.referral_status,
                "referral_place": v.referral_place,
                "control_date": v.control_date,
                "control_done": v.control_done,
            }
            for v in db.query(UKSVisitORM).all()
        ],
        "medicines": [
            {"name": m.name, "unit": m.unit, "stock": m.stock, "minimum_stock": m.minimum_stock}
            for m in db.query(MedicineInventoryORM).all()
        ],
    }
    data = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    return StreamingResponse(
        BytesIO(data),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="backup_emr_uks_{date.today().isoformat()}.json"'},
    )


@router.post("/admin/restore")
def restore_backup(
    payload: dict,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles(ROLE_ADMIN)),
):
    restored = {"patients": 0, "visits": 0, "medicines": 0}
    for item in payload.get("patients", []):
        patient = db.get(PatientORM, str(item.get("id")))
        if patient is None:
            patient = PatientORM(id=str(item.get("id")), name=item.get("name") or "-", age=int(item.get("age") or 0), gender=item.get("gender") or "-")
            db.add(patient)
        patient.name = item.get("name") or patient.name
        patient.age = int(item.get("age") or patient.age or 0)
        patient.gender = item.get("gender") or patient.gender
        patient.class_name = item.get("class_name")
        patient.birth_date = item.get("birth_date")
        patient.parent_name = item.get("parent_name")
        patient.parent_phone = item.get("parent_phone")
        restored["patients"] += 1

    for item in payload.get("medicines", []):
        med = db.query(MedicineInventoryORM).filter(MedicineInventoryORM.name == item.get("name")).first()
        if med is None:
            med = MedicineInventoryORM(name=item.get("name") or "-", unit=item.get("unit") or "tablet", stock=0, minimum_stock=10)
            db.add(med)
        med.unit = item.get("unit") or med.unit
        med.stock = int(item.get("stock") or 0)
        med.minimum_stock = int(item.get("minimum_stock") or 0)
        restored["medicines"] += 1

    for item in payload.get("visits", []):
        patient_id = str(item.get("patient_id") or "")
        if not patient_id or db.get(PatientORM, patient_id) is None:
            continue
        visit = None
        if item.get("id") is not None:
            visit = db.get(UKSVisitORM, int(item.get("id")))
        if visit is None:
            visit = UKSVisitORM(
                patient_id=patient_id,
                visit_date=item.get("visit_date") or date.today().isoformat(),
                complaint=item.get("complaint") or "-",
                examination=item.get("examination") or "-",
                treatment=item.get("treatment") or "-",
            )
            db.add(visit)
        visit.patient_id = patient_id
        visit.visit_date = item.get("visit_date") or visit.visit_date
        visit.complaint = item.get("complaint") or visit.complaint
        visit.examination = item.get("examination") or visit.examination
        visit.treatment = item.get("treatment") or visit.treatment
        visit.diagnosis = item.get("diagnosis")
        visit.notes = item.get("notes")
        visit.referral_to = item.get("referral_to")
        visit.referral_status = item.get("referral_status")
        visit.referral_place = item.get("referral_place")
        visit.control_date = item.get("control_date")
        visit.control_done = bool(item.get("control_done"))
        restored["visits"] += 1
    write_audit_log(db, current_user, "restore_backup", "system", None, json.dumps(restored))
    db.commit()
    return {"message": "Restore selesai", "restored": restored}


@router.post("/patients/import-excel")
def import_patients_excel(
    payload: dict,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles(ROLE_ADMIN)),
):
    content = payload.get("content_base64")
    if not content:
        raise HTTPException(status_code=400, detail="content_base64 wajib diisi")
    raw = base64.b64decode(content.split(",", 1)[-1])
    workbook = load_workbook(BytesIO(raw), data_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip().lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    aliases = {
        "id": ["id", "nis", "id / nis"],
        "name": ["nama", "nama lengkap", "name", "full name"],
        "gender": ["gender", "jenis kelamin", "jk"],
        "birth_date": ["tanggal lahir", "birth date", "birth_date"],
        "class_name": ["kelas", "class", "class_name"],
        "parent_name": ["wali asuh", "nama wali asuh", "orang tua", "parent name"],
        "parent_phone": ["nomor hp wali asuh", "no hp", "hp wali", "parent phone"],
    }

    def idx(key: str) -> int | None:
        for alias in aliases[key]:
            if alias in headers:
                return headers.index(alias)
        return None

    id_idx = idx("id")
    name_idx = idx("name")
    if id_idx is None or name_idx is None:
        raise HTTPException(status_code=400, detail="Kolom NIS/ID dan Nama wajib ada")

    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[id_idx] or not row[name_idx]:
            continue
        item = {
            "id": str(row[id_idx]).strip(),
            "name": str(row[name_idx]).strip(),
            "gender": str(row[idx("gender")] or "-").strip() if idx("gender") is not None else "-",
            "birth_date": str(row[idx("birth_date")] or "").strip() if idx("birth_date") is not None else None,
            "class_name": str(row[idx("class_name")] or "").strip() if idx("class_name") is not None else None,
            "parent_name": str(row[idx("parent_name")] or "").strip() if idx("parent_name") is not None else None,
            "parent_phone": str(row[idx("parent_phone")] or "").strip() if idx("parent_phone") is not None else None,
        }
        rows.append(item)

    if payload.get("preview", False):
        return {"preview": rows[:20], "total": len(rows)}

    created = 0
    updated = 0
    for item in rows:
        patient = db.get(PatientORM, item["id"])
        if patient is None:
            patient = PatientORM(id=item["id"], name=item["name"], age=0, gender=item["gender"])
            db.add(patient)
            created += 1
        else:
            updated += 1
        patient.name = item["name"]
        patient.gender = item["gender"]
        patient.class_name = item["class_name"]
        patient.birth_date = item["birth_date"]
        patient.parent_name = item["parent_name"]
        patient.parent_phone = item["parent_phone"]
    write_audit_log(db, current_user, "import_patients_excel", "patient", None, f"created={created}; updated={updated}")
    db.commit()
    return {"message": "Import siswa selesai", "created": created, "updated": updated, "total": len(rows)}


@router.post("/uks/visits/{visit_id}/notify-rest-letter")
def notify_rest_letter(
    visit_id: int,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles(ROLE_ADMIN, ROLE_PERAWAT)),
):
    visit = db.get(UKSVisitORM, visit_id)
    if visit is None:
        raise HTTPException(status_code=404, detail="UKS visit not found")
    patient = db.get(PatientORM, visit.patient_id)
    if patient is None:
        raise HTTPException(status_code=404, detail="Patient not found")
    status_text, message = send_whatsapp_message(patient.parent_phone, build_rest_letter_whatsapp_message(patient, visit))
    write_audit_log(db, current_user, "notify_rest_letter", "uks_visit", visit.id, f"{status_text}: {message}")
    db.commit()
    return {"whatsapp_status": status_text, "whatsapp_message": message}


def _visit_or_404(db: Session, visit_id: int) -> tuple[UKSVisitORM, PatientORM]:
    visit = db.get(UKSVisitORM, visit_id)
    if visit is None:
        raise HTTPException(status_code=404, detail="UKS visit not found")
    patient = db.get(PatientORM, visit.patient_id)
    if patient is None:
        raise HTTPException(status_code=404, detail="Patient not found")
    return visit, patient


@router.get("/uks/visits/{visit_id}/referral-letter")
def uks_referral_letter_pdf(
    visit_id: int,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles(ROLE_ADMIN, ROLE_PERAWAT)),
):
    visit, patient = _visit_or_404(db, visit_id)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=24, bottomMargin=28, leftMargin=42, rightMargin=42)
    styles = getSampleStyleSheet()
    elements = []
    _append_pdf_letterhead(elements, doc, "SURAT RUJUKAN UKS", f"Tanggal: {visit.visit_date}", styles)
    rows = [
        ["Nama", patient.name],
        ["NIS", patient.id],
        ["Kelas", patient.class_name or "-"],
        ["Keluhan", visit.complaint],
        ["Diagnosa", visit.diagnosis or "-"],
        ["Tindakan", visit.treatment],
        ["Tujuan Rujukan", visit.referral_to or visit.referral_place or "Fasilitas kesehatan terdekat"],
    ]
    elements.append(Table(rows, colWidths=[130, doc.width - 130], style=TableStyle([
        ("GRID", (0, 0), (-1, -1), .4, colors.lightgrey),
        ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("PADDING", (0, 0), (-1, -1), 7),
    ])))
    _append_pdf_signature(elements, doc, current_user, styles, "Petugas UKS")
    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="surat_rujukan_{visit_id}.pdf"'})


@router.get("/uks/visits/{visit_id}/rest-letter")
def uks_rest_letter_pdf(
    visit_id: int,
    reason: str = Query(default="Istirahat"),
    days: int = Query(default=1, ge=1, le=30),
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles(ROLE_ADMIN, ROLE_PERAWAT)),
):
    visit, patient = _visit_or_404(db, visit_id)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=24, bottomMargin=28, leftMargin=42, rightMargin=42)
    styles = getSampleStyleSheet()
    elements = []
    _append_pdf_letterhead(elements, doc, "SURAT IZIN ISTIRAHAT UKS", f"Tanggal: {visit.visit_date}", styles)
    rows = [
        ["Nama", patient.name],
        ["NIS", patient.id],
        ["Kelas", patient.class_name or "-"],
        ["Alasan Izin", reason],
        ["Lama Istirahat", f"{days} hari"],
        ["Diagnosa", visit.diagnosis or "-"],
        ["Catatan", "Disarankan istirahat dan pemantauan kondisi oleh wali asuh/orang tua."],
    ]
    elements.append(Table(rows, colWidths=[130, doc.width - 130], style=TableStyle([
        ("GRID", (0, 0), (-1, -1), .4, colors.lightgrey),
        ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("PADDING", (0, 0), (-1, -1), 7),
    ])))
    _append_pdf_signature(elements, doc, current_user, styles, "Petugas UKS")
    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="surat_izin_{visit_id}.pdf"'})


@router.post("/ckg/students/{student_id}/notify-completed")
def notify_ckg_completed(
    student_id: int,
    db: Session = Depends(get_db),
    current_user: UserORM = Depends(require_roles(ROLE_ADMIN, ROLE_PERAWAT)),
):
    student = db.get(CKGStudentORM, student_id)
    if student is None:
        raise HTTPException(status_code=404, detail="CKG student not found")
    patient = db.get(PatientORM, student.nis)
    phone = student.parent_phone or (patient.parent_phone if patient else None)
    parent_name = student.parent_name or (patient.parent_name if patient else None) or "Wali Asuh / Orang Tua"
    message = f"""[UKS SRMA 13 Bekasi]

Yth. {parent_name},

Hasil CKG siswa {student.full_name} telah selesai diproses.

Status: {student.status}
Silakan hubungi petugas UKS bila diperlukan tindak lanjut."""
    status_text, detail = send_whatsapp_message(phone, message)
    write_audit_log(db, current_user, "notify_ckg_completed", "ckg_student", student.id, f"{status_text}: {detail}")
    db.commit()
    return {"whatsapp_status": status_text, "whatsapp_message": detail}


@router.get("/audit-logs/export/excel")
def export_audit_logs_excel(
    db: Session = Depends(get_db),
    _: UserORM = Depends(require_roles(ROLE_ADMIN)),
):
    logs = db.query(AuditLogORM).order_by(AuditLogORM.timestamp.desc()).limit(5000).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Log"
    ws.append(["Timestamp", "User", "Action", "Entity", "Entity ID", "Details"])
    for log in logs:
        ws.append([str(log.timestamp), log.username, log.action, log.entity_type, log.entity_id, log.details])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="audit_log.xlsx"'},
    )
    PasswordResetRequest,

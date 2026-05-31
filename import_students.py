from openpyxl import load_workbook

from app.db.database import SessionLocal
from app.db.models import PatientORM


def main() -> None:
    db = SessionLocal()
    try:
        wb = load_workbook("students_clean_import.xlsx", data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print("File kosong.")
            return

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        idx = {name: i for i, name in enumerate(headers)}

        required = ["id", "name", "gender", "class_name", "birth_date"]
        missing = [col for col in required if col not in idx]
        if missing:
            print(f"Kolom wajib tidak ditemukan: {', '.join(missing)}")
            return

        inserted = 0
        updated = 0

        for row in rows[1:]:
            student_id = str(row[idx["id"]]).strip() if row[idx["id"]] is not None else ""
            if not student_id:
                continue

            name = str(row[idx["name"]]).strip() if row[idx["name"]] is not None else ""
            gender = str(row[idx["gender"]]).strip() if row[idx["gender"]] is not None else "L"
            class_name = str(row[idx["class_name"]]).strip() if row[idx["class_name"]] is not None else None
            birth_date = str(row[idx["birth_date"]]).strip() if row[idx["birth_date"]] is not None else None

            existing = db.get(PatientORM, student_id)
            if existing:
                existing.name = name
                existing.gender = gender
                existing.class_name = class_name
                existing.birth_date = birth_date
                db.add(existing)
                updated += 1
            else:
                db.add(
                    PatientORM(
                        id=student_id,
                        name=name,
                        gender=gender,
                        class_name=class_name,
                        birth_date=birth_date,
                        age=0,
                    )
                )
                inserted += 1

        db.commit()
        print(f"Import siswa berhasil. Ditambah: {inserted}, diupdate: {updated}.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
